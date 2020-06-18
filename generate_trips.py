from trip import Trip, load_from_row
from openpyxl import load_workbook
from geojson import Feature, Point, FeatureCollection, dump


def generate_geojson(file: str, dest: str):
    wb = load_workbook(filename=file)
    features = []
    for sheet in wb.sheetnames:
        max_row = wb[sheet].max_row
        max_col = wb[sheet].max_column
        for row in wb[sheet].iter_rows(min_row=2, min_col=1, max_col=max_col, max_row=max_row):
            trip = load_from_row(row)
            trip.print()
            if trip.start_check is not None:
                trip.start_check = trip.start_check.split(",")[0]
                point = Point((trip.start_lon, trip.start_lat))

                props = trip.__dict__
                props["date"] = props["date"].strftime("%m/%d/%Y")
                features.append(Feature(geometry=point, properties=props))

    feature_collection = FeatureCollection(features)
    with open(dest, 'w') as outfile:
        outfile.write("var journal_bord = ")
        dump(feature_collection, outfile, sort_keys=False)


if __name__ == '__main__':  # pragma: no cover
    generate_geojson("./journaux de bord adapt main.xlsx", "./web/journal_bord.geojson.js")
