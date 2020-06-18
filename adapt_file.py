from openpyxl import load_workbook, Workbook
from datetime import datetime
from typing import Tuple
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
from time import sleep
from trip import Trip

titles = ["Bateau", "Date", "Départ raw", "Arrivée raw", "Départ", "Arrivée", "Lat. Départ", "Long. Départ", "Lat. Arrivée",
          "Long. Arrivée", "Commentaire"]


def get_trip(val: str) -> Tuple[str, str]:
    if val.count("_") > 0:
        splt = val.split("_")
    else:
        splt = val.split("-")
    if len(splt) == 0:
        raise ValueError("Error in trip value")
    elif len(splt) == 1:
        return [(splt[0],)]
    elif len(splt) == 2:
        return [(splt[0], splt[1])]
    elif len(splt) == 3:
        return [(splt[0], splt[1]), (splt[1], splt[2])]
    else:
        raise ValueError("Trip contains more than 3 places")


def adapt_file(file: str, dest: str):
    wb = load_workbook(filename=file)
    ws = Workbook()
    for i_col, title in enumerate(titles):
        ws.active.cell(row=1, column=i_col + 1).value = title
    curr_row = 2
    for sheet in wb.sheetnames:
        max_row = wb[sheet].max_row
        max_col = wb[sheet].max_column
        for i_col, col in enumerate(wb[sheet].iter_cols(min_row=2, min_col=2,
                                                        max_col=max_col, max_row=max_row)):
            for i_row, cell in enumerate(col):
                if cell.value is not None:
                    date = wb[sheet].cell(row=1, column=i_col + 2).value
                    date_with_day = datetime(year=date.year, month=date.month, day=i_row + 1)
                    places = get_trip(cell.value)

                    for place in places:
                        current_trip = Trip(boat=sheet, date=date_with_day, start=place[0])
                        if len(place) > 1:
                            current_trip.end = place[1]
                        if cell.comment is not None:
                            current_trip.comment = cell.comment.text
                        current_trip.write_to_ws(ws, row=curr_row)
                        curr_row += 1
    ws.save(dest)


def get_geo(pos: str, locator: Nominatim):
    return locator.geocode(pos, country_codes=['gb', 'fr', 'es', 'pt'])


def get_place_from_raw(row, locator: Nominatim):
    if row[4].value is None and row[2].value is not None:
        start_pos = get_geo(row[2].value, locator)
        sleep(0.5)
        if start_pos is not None:
            print(f"Found matching [{row[2].value}] - > [{start_pos.address}]")
            row[4].value = start_pos.address
            row[5].value = start_pos.latitude
            row[6].value = start_pos.longitude
    if row[7].value is None and row[3].value is not None:
        end_pos = get_geo(row[3].value, locator)
        sleep(0.5)
        if end_pos is not None:
            print(f"Found matching [{row[3].value}] - > [{end_pos.address}]")
            row[7].value = end_pos.address
            row[8].value = end_pos.latitude
            row[9].value = end_pos.longitude


def add_positions(file: str):
    locator = Nominatim(user_agent="osm")
    wb = load_workbook(filename=file)
    for sheet in wb.sheetnames:
        max_row = wb[sheet].max_row
        max_col = wb[sheet].max_column
        for row in wb[sheet].iter_rows(min_row=0, min_col=1, max_col=max_col, max_row=max_row):
            try:
                get_place_from_raw(row, locator)
            except GeocoderTimedOut:
                print(f"Geocoder time out at row {row[1].value}")
                break

    wb.save(file)


if __name__ == '__main__':  # pragma: no cover
    adapt_file("./journaux de bord.xlsx", "./journaux de bord adapt.xlsx")
    add_positions("./journaux de bord adapt main.xlsx")