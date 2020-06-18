from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import numbers
from typing import Tuple


class Trip(object):
    def __init__(self, boat: str, date: datetime, start: str = '', end: str = '',
                 comment: str = ''):
        """ A structure for trip management"""
        self.boat = boat
        """Trip date as a datetime"""
        self.date = date
        """Start location"""
        self.start = start
        """End location"""
        self.end = end
        """Start checked location"""
        self.start_check = None
        """End checked location"""
        self.end_check = None
        """Start latitude"""
        self.start_lat = None
        """Start longitude"""
        self.start_lon = None
        """End latitude"""
        self.end_lat = None
        """End longitude"""
        self.end_lon = None
        """Comment"""
        self.comment = comment

    def print(self):
        res = f"{self.boat} le {self.date}"
        if self.start_check is not None:
            res += f" au départ de {self.start_check.split(',')[0]}"
        if self.end_check is not None:
            res += f" à destination de {self.end_check.split(',')[0]}"
        res += f": {self.comment}"
        print(res)

    def write_to_ws(self, ws: Workbook, row: int):
        ws.active.cell(row=row, column=1).value = self.boat
        ws.active.cell(row=row, column=2).number_format = \
            numbers.FORMAT_DATE_DDMMYY
        ws.active.cell(row=row, column=2).value = self.date
        ws.active.cell(row=row, column=3).value = self.start
        ws.active.cell(row=row, column=4).value = self.end
        ws.active.cell(row=row, column=5).value = self.start_check
        ws.active.cell(row=row, column=6).value = self.end_check
        ws.active.cell(row=row, column=7).value = self.start_lat
        ws.active.cell(row=row, column=8).value = self.start_lon
        ws.active.cell(row=row, column=9).value = self.end_lat
        ws.active.cell(row=row, column=10).value = self.end_lon
        ws.active.cell(row=row, column=11).value = self.comment


def load_from_row(row: Tuple) -> Trip:
    trip = Trip(row[0].value, row[1].value, row[2].value, row[3].value, row[10].value)
    trip.start_check = row[4].value
    trip.end_check = row[7].value
    trip.start_lat = row[5].value
    trip.start_lon = row[6].value
    trip.end_lat = row[8].value
    trip.end_lon = row[9].value
    return trip
